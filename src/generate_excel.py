import xlrd
import xlwt
from openpyxl import load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    Series,
)

# Create a new workbook and worksheet
workbook = xlwt.Workbook()
worksheet = workbook.add_sheet('Sheet1')

# Add value data
worksheet.write(0, 0, 'Date')
worksheet.write(0, 1, 'Value')
worksheet.write(1, 0, '1/1/2021')
worksheet.write(1, 1, 100)
worksheet.write(2, 0, '1/2/2021')
worksheet.write(2, 1, 200)
worksheet.write(3, 0, '1/3/2021')
worksheet.write(3, 1, 300)

# Save the workbook to disk
workbook.save('example.xls')

# Open the workbook with openpyxl
book = load_workbook('example.xls')
sheet = book.active

# Create a new chart and add data
chart = LineChart()
values = Reference(sheet, min_col=2, min_row=1, max_row=3)
categories = Reference(sheet, min_col=1, min_row=2, max_row=3)
series = Series(values, title='Values')
chart.add_data(series, titles_from_data=True)
chart.set_categories(categories)
chart.title = 'Example Chart'
chart.x_axis.title = 'Date'
chart.y_axis.title = 'Value'

# Add the chart to the worksheet
sheet.add_chart(chart, 'D3')

# Save changes and output file
book.save('example.xls')
